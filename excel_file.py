import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from aiogram import Bot, Dispatcher, types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor

BOT_TOKEN = "7332023971:AAFOYSIBfUH0PErGWhADpZKsZyz0LMcQm5E"

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(bot)

EXCEL_FILE = "tasks.xlsx"

RED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
GREEN_FILL = PatternFill(start_color="66CC66", end_color="66CC66", fill_type="solid")


students = []

def init_excel():
    try:
        load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        df = pd.DataFrame(columns=["Name"] + [str(i) for i in range(1, 32)])
        df["Name"] = ["Shamsiddin", "Alisher", "Baxtiyor", "Diyas", "Behruz", "Sardor", "Sarvar"]
        df.to_excel(EXCEL_FILE, index=False)
        print("Yangi Excel File yaratildi! .")

def load_students():
    global students
    df = pd.read_excel(EXCEL_FILE)
    students = df["Name"].tolist()

# def add_task(student_name, task_text):
#     today = datetime.now().day
#     df = pd.read_excel(EXCEL_FILE)
#     row_index = df.index[df["Name"] == student_name].tolist()[0]
#     col_name = str(today)
#
#     df.loc[row_index, col_name] = task_text
#     df.to_excel(EXCEL_FILE, index=False)
#
#     workbook = load_workbook(EXCEL_FILE)
#     sheet = workbook.active
#     cell = sheet.cell(row=row_index + 2, column=today + 1)
#     cell.font = Font(color="9C0006", bold=True)
#     sheet.column_dimensions[cell].width = 20
#     cell.fill = RED_FILL
#     workbook.save(EXCEL_FILE)

def add_task(student_name, task_text):
    today = datetime.now().day
    df = pd.read_excel(EXCEL_FILE)

    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    df.loc[row_index, col_name] = task_text
    df.to_excel(EXCEL_FILE, index=False)

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    cell = sheet.cell(row=row_index + 2, column=today + 1)
    cell.fill = RED_FILL
    uzunligi = len(task_text)

    sheet.column_dimensions[sheet.cell(row=1, column=today + 1).column_letter].width = uzunligi

    workbook.save(EXCEL_FILE)


def approve_task(student_name):
    today = datetime.now().day
    df = pd.read_excel(EXCEL_FILE)
    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    if pd.notna(df.loc[row_index, col_name]):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        cell = sheet.cell(row=row_index + 2, column=today + 1)
        cell.font = Font(color="006100", bold=True)
        cell.fill = GREEN_FILL
        workbook.save(EXCEL_FILE)

def get_active_tasks():
    today = datetime.now().day
    df = pd.read_excel(EXCEL_FILE)
    active_students = []
    for _, row in df.iterrows():
        if pd.isna(row[str(today)]):
            active_students.append(row["Name"])
    return active_students




@dp.message_handler(commands=["start"])
async def start_command(message: types.Message):
    keyboard = InlineKeyboardMarkup()
    for student in students:
        keyboard.add(InlineKeyboardButton(text=student, callback_data=f"student_{student}"))
    await message.reply("Oquvchini tanlang:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("student_"))
async def student_selected(callback_query: types.CallbackQuery):
    student_name = callback_query.data.split("_")[1]
    today = datetime.now().day
    df = pd.read_excel(EXCEL_FILE)

    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    if pd.isna(df.loc[row_index, col_name]):
        await bot.send_message(callback_query.from_user.id, f"Oquvchiga vazifa yuboring {student_name}:")
        @dp.message_handler()
        async def input_task(message: types.Message):
            add_task(student_name, message.text)
            await message.reply(f"Vazifa {student_name} uchun berildi.")
    else:
        keyboard = InlineKeyboardMarkup()
        keyboard.add(InlineKeyboardButton(text="Tasdiqlash", callback_data=f"approve_{student_name}"))
        await bot.send_message(callback_query.from_user.id, f"{student_name} vazifasi allaqachon bor: {df.loc[row_index, col_name]}", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith("approve_"))
async def approve_selected(callback_query: types.CallbackQuery):
    student_name = callback_query.data.split("_")[1]
    approve_task(student_name)
    await bot.send_message(callback_query.from_user.id, f"{student_name} uchun vazifa tasdiqlandi.")

@dp.message_handler(commands=["active"])
async def active_command(message: types.Message):
    task_yolar = get_active_tasks()
    if task_yolar:
        await message.reply("Bosh oquvchilar:\n" + "\n".join(task_yolar))
    else:
        await message.reply("Bosh oquvchilar yoq!")






if __name__ == "__main__":
    init_excel()
    load_students()
    executor.start_polling(dp, skip_updates=True)
