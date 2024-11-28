from aiogram import Bot, Dispatcher, types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor
from aiogram.dispatcher import FSMContext
from aiogram.contrib.fsm_storage.memory import MemoryStorage

import pandas as pd
from openpyxl import load_workbook, workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

BOT_TOKEN = "TOKEN"

bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

month_name = datetime.now().strftime("%B")

EXCEL_FILE = "tasks.xlsx"

RED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
GREEN_FILL = PatternFill(start_color="66CC66", end_color="66CC66", fill_type="solid")

students = []


def init_excel():
    try:
        load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        df = pd.DataFrame(columns=["Name"] + [str(i) for i in range(1, 32)])
        df["Name"] = ["Shamsiddin", "Alisher", "Baxtiyor", "Diyas", "Behruz", "Sardor", "Sarvar"]
        df.to_excel(EXCEL_FILE, index=False)




def load_students():
    global students
    df = pd.read_excel(EXCEL_FILE)
    students = df["Name"].tolist()


def add_task(student_name, task_text):
    today = datetime.now().day

    df = pd.read_excel(EXCEL_FILE)

    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    df.loc[row_index, col_name] = task_text
    df.to_excel(EXCEL_FILE, index=False)

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        for col in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value and cell.fill != GREEN_FILL:
                cell.fill = RED_FILL

    new_task_cell = sheet.cell(row=row_index + 2, column=today + 1)
    new_task_cell.value = task_text
    new_task_cell.fill = RED_FILL

    column_letter = sheet.cell(row=1, column=today + 1).column_letter
    sheet.column_dimensions[column_letter].width = max(20, len(task_text))

    workbook.save(EXCEL_FILE)


def approve_task(student_name):
    today = datetime.now().day
    df = pd.read_excel(EXCEL_FILE)
    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    if pd.notna(df.loc[row_index, col_name]):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        cell = sheet.cell(row=row_index + 2, column=today + 1)
        cell.font = Font(color="006100", bold=True)
        cell.fill = GREEN_FILL
        workbook.save(EXCEL_FILE)


@dp.message_handler(commands=["start"])
async def start_command(message: types.Message):
    keyboard = InlineKeyboardMarkup()
    for student in students:
        keyboard.add(InlineKeyboardButton(text=student, callback_data=f"student_{student}"))
    await message.reply("Assalomu Aleykum\nOquvchilardan birini tanlang:", reply_markup=keyboard)


@dp.callback_query_handler(lambda c: c.data.startswith("student_"))
async def student_selected(callback_query: types.CallbackQuery, state: FSMContext):
    student_name = callback_query.data.split("_")[1]
    today = datetime.now().day
    df = pd.read_excel(EXCEL_FILE)

    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    if pd.isna(df.loc[row_index, col_name]):
        await state.update_data(selected_student=student_name)
        await bot.send_message(callback_query.from_user.id, f"{student_name} uchun vazifa kiriting:")
    else:
        keyboard = InlineKeyboardMarkup()
        keyboard.add(InlineKeyboardButton(text="Tasdiqlash", callback_data=f"approve_{student_name}"))
        # keyboard.add(InlineKeyboardButton(text="Yana vazifa berish", callback_data=f"new_task_{student_name}"))
        await bot.send_message(callback_query.from_user.id,
                               f"{student_name} uchun vazifa mavjud: {df.loc[row_index, col_name]}",
                               reply_markup=keyboard)


@dp.callback_query_handler(lambda c: c.data.startswith("approve_"))
async def approve_selected(callback_query: types.CallbackQuery):
    student_name = callback_query.data.split("_")[1]
    approve_task(student_name)
    # keyboard = InlineKeyboardMarkup()
    # keyboard.add(InlineKeyboardButton(text="Яна вазифа бериш", callback_data=f"new_task_{student_name}"))
    await bot.send_message(callback_query.from_user.id, f"{student_name} uchun vazufa tasdiqlandi.")


# @dp.callback_query_handler(lambda c: c.data.startswith("new_task_"))
# async def new_task(callback_query: types.CallbackQuery, state: FSMContext):
#     student_name = callback_query.data.split("_")[1]
#     await state.update_data(selected_student=student_name)
#     await bot.send_message(callback_query.from_user.id, f"{student_name} учун янги вазифани киритинг:")


@dp.message_handler(state="*", content_types=types.ContentType.TEXT)
async def input_task(message: types.Message, state: FSMContext):
    data = await state.get_data()
    student_name = data.get("selected_student")
    if student_name:
        add_task(student_name, message.text)
        await message.reply(f"Vazifa {student_name} uchun berildi.")
        await state.finish()


if __name__ == "__main__":
    init_excel()
    load_students()
    executor.start_polling(dp, skip_updates=True)
