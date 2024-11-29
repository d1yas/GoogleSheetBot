from math import trunc

from aiogram import Bot, Dispatcher, types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor
from aiogram.dispatcher import FSMContext
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from states.state import Main
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

BOT_TOKEN = "sss"

bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

EXCEL_FILE = "tasks.xlsx"
ADMIN = 123456

RED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
GREEN_FILL = PatternFill(start_color="66CC66", end_color="66CC66", fill_type="solid")

students = []


def get_active_tasks():
    today = datetime.now().day
    df = pd.read_excel(EXCEL_FILE)
    active_students = []
    for _, row in df.iterrows():
        if pd.isna(row[str(today)]):
            active_students.append(row["Name"])
    return active_students

def init_excel():
    try:
        workbook = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(EXCEL_FILE)

    workbook = load_workbook(EXCEL_FILE)
    current_month = datetime.now().strftime("%B")

    if current_month not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=current_month)
        sheet.append(["Name"] + [str(i) for i in range(1, 32)])  # Добавляем заголовки
        students_data = [
            "Shamsiddin", "Alisher", "Baxtiyor", "Diyas", "Behruz", "Sardor", "Sarvar"
        ]
        for student in students_data:
            sheet.append([student] + [""] * 31)
        workbook.save(EXCEL_FILE)


def load_students():
    global students
    workbook = load_workbook(EXCEL_FILE)
    current_month = datetime.now().strftime("%B")
    if current_month not in workbook.sheetnames:
        init_excel()
    df = pd.DataFrame(workbook[current_month].values)
    students = df.iloc[1:, 0].tolist()


def add_task(student_name, task_text):
    today = datetime.now().day
    current_month = datetime.now().strftime("%B")

    workbook = load_workbook(EXCEL_FILE)

    if current_month not in workbook.sheetnames:
        init_excel()
    sheet = workbook[current_month]

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == student_name:
            task_cell = sheet.cell(row=row, column=today + 1)
            task_cell.value = task_text
            task_cell.fill = RED_FILL

            column_letter = sheet.cell(row=1, column=today + 1).column_letter
            sheet.column_dimensions[column_letter].width = max(
                sheet.column_dimensions[column_letter].width, len(task_text) + 5
            )
            break

    workbook.save(EXCEL_FILE)



def approve_task(student_name):

    today = datetime.now().day
    current_month = datetime.now().strftime("%B")

    workbook = load_workbook(EXCEL_FILE)

    if current_month not in workbook.sheetnames:
        init_excel()

    sheet = workbook[current_month]
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]
    df = df[1:]

    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    if pd.notna(df.loc[row_index, col_name]):
        cell = sheet.cell(row=row_index + 1, column=today+1)
        cell.font = Font(color="FF006100", bold=True)
        cell.fill = GREEN_FILL
        workbook.save(EXCEL_FILE)


@dp.message_handler(commands=["start"])
async def start_command(message: types.Message):
    if message.from_user.id == ADMIN:
        keyboard = InlineKeyboardMarkup()
        for student in students:
            keyboard.add(InlineKeyboardButton(text=student, callback_data=f"student_{student}"))
        await message.reply(f"👋Assalomu Aleykum - <b>{message.from_user.full_name}</b>\n👨‍🎓O'quvchilardan birini tanlang:", reply_markup=keyboard, parse_mode="HTML")
    else:
        await message.answer("Siz bu botni ishlatish huquqiga ega emassiz❌")



@dp.callback_query_handler(lambda c: c.data.startswith("student_"))
async def student_selected(callback_query: types.CallbackQuery, state: FSMContext):
    student_name = callback_query.data.split("_")[1]
    today = datetime.now().day
    current_month = datetime.now().strftime("%B")

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook[current_month]
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]
    df = df[1:]

    row_index = df.index[df["Name"] == student_name].tolist()[0]
    col_name = str(today)

    cell = sheet.cell(row=row_index + 1, column=today + 1)

    if pd.isna(df.loc[row_index, col_name]):
        await state.update_data(selected_student=student_name)
        await bot.send_message(callback_query.from_user.id, f"📝{student_name} uchun vazifa kiriting:")
        await Main.student_state_name.set()

    else:
        if cell.font == Font(color="FF006100", bold=True):
            await bot.send_message(callback_query.from_user.id, f"📕{student_name} vazifasi bajarilgan.")
        else:
            keyboard = InlineKeyboardMarkup()
            keyboard.add(InlineKeyboardButton(text="Tasdiqlash", callback_data=f"approve_{student_name}"))
            await bot.send_message(
                callback_query.from_user.id,
                f"✔️{student_name} uchun vazifa mavjud: {df.loc[row_index, col_name]}",
                reply_markup=keyboard
            )
@dp.callback_query_handler(lambda c: c.data.startswith("approve_"))
async def approve_selected(callback_query: types.CallbackQuery):
    student_name = callback_query.data.split("_")[1]
    approve_task(student_name)
    await bot.send_message(callback_query.from_user.id, f"✅{student_name} uchun vazifa tasdiqlandi.")


@dp.message_handler(state=Main.student_state_name, content_types=types.ContentType.TEXT)
async def input_task(message: types.Message, state: FSMContext):
    data = await state.get_data()
    student_name = data.get("selected_student")
    if student_name:
        add_task(student_name, message.text)
        await message.reply(f"✋Vazifa {student_name} uchun berildi.")
        await state.finish()


@dp.message_handler(commands=["getxlsx"])
async def get_xlsx(message: types.Message):
    if message.from_user.id == ADMIN:
        await bot.send_document(message.from_user.id, open(EXCEL_FILE, "rb"))
    else:
        await message.answer("Siz bizning exel fileni olish huquqiga ega emassiz❌")


@dp.message_handler(commands=['statistic'])
async def send_statistics(message: types.Message):
    if message.from_user.id == ADMIN:
        global count_student
        today = datetime.now().day
        current_month = datetime.now().strftime("%B")

        try:
            workbook = load_workbook(EXCEL_FILE)
            if current_month not in workbook.sheetnames:
                await message.reply("Joriy oy uchun ma'lumotlar topilmadi.")
                return

            sheet = workbook[current_month]
            completed_students = []

            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=today + 1):
                student_name = row[0].value
                cell = row[today]
                if cell.font == Font(color="FF006100", bold=True):
                    completed_students.append("✅ "+student_name)

            if completed_students:
                a= len(completed_students)
                if a == 7:
                    count_student = "🏆Bugun barcha O'quvchilaringiz vazifa bajargan"
                elif a < 7:
                    not_completed_count_students = 0
                    for i in range(a):
                        not_completed_count_students += 1
                    which_student_not_completed_count = 7 - not_completed_count_students
                    count_student = f"❌Bugun {which_student_not_completed_count}ta o'quvchi vazifa qilmagan"

                response = f"{count_student}\n\n🕔<b>Bugun vazifani bajargan o'quvchilar:</b>\n\n" + "\n".join(completed_students)
            else:
                response = "Bugun hech bir o'quvchi vazifani bajarmagan."

            await message.reply(response, parse_mode=types.ParseMode.HTML)

        except FileNotFoundError:
            await message.reply("Excel fayli topilmadi.")
        except Exception as e:
            await message.reply(f"Xatolik yuz berdi: {e}")
    else:
        await message.answer("Siz bu botda admin emassiz❌")

if __name__ == "__main__":
    init_excel()
    load_students()
    executor.start_polling(dp, skip_updates=True)
